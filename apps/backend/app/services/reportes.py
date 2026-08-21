"""Reporte mensual en Excel para que los bodegueros le manden a sus
superiores -- a nivel de documento (entrega), no de producto: fecha/tipo/
numero/cantidad son atributos consolidados de la entrega (un documento
puede traer varios productos, la cantidad y el historial de ocasiones van
sumados/consolidados, no desglosados por producto).
"""

import io
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.db import get_pool

_MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# Los 4 tipos conocidos van primero en cada hoja; cualquier otro (custom, ver
# vision.py/EntregaRevision) va despues, en orden alfabetico.
_ORDEN_TIPOS_CONOCIDOS = ("FEI", "TB", "RM3", "RM2")

# Mismos valores que MotivoDevolucion/ResolucionDevolucion en
# app/models/devolucion.py -- version legible para el reporte.
_MOTIVOS_LEGIBLES = {
    "danado": "Dañado",
    "equivocado": "Equivocado",
    "vencido": "Vencido",
    "no_era_lo_pedido": "No era lo pedido",
    "otro": "Otro",
}
_RESOLUCIONES_LEGIBLES = {
    "reposicion": "Reposición",
    "reembolso": "Reembolso",
}


def _orden_tipo(tipo: str) -> tuple[int, str]:
    if tipo in _ORDEN_TIPOS_CONOCIDOS:
        return (_ORDEN_TIPOS_CONOCIDOS.index(tipo), tipo)
    return (len(_ORDEN_TIPOS_CONOCIDOS), tipo)


def _ocasiones_por_entrega(
    logs_rows: list,
) -> dict[str, list[tuple[datetime, int]]]:
    """A partir de los logs 'entrega_actualizada' (que traen la foto de
    cantidad_entregada de cada item en ese momento -- ver
    aplicar_actualizacion_items en duplicates.py), reconstruye cuantas
    ocasiones reales de entrega tuvo cada documento, con fecha y cantidad
    total (sumada entre productos) de cada una.

    La primera vez que aparece un item cuenta como su primera ocasion (el
    delta es el valor completo, no contra 0 previo real sino contra "nunca
    visto"). Un delta <= 0 (una devolucion, o un guardado que solo cambio
    una nota sin tocar cantidades) no cuenta como ocasion de entrega.
    """
    # entrega_id -> item_id -> ultimo cantidad_entregada conocido
    ultimo_conocido: dict[str, dict[str, int]] = defaultdict(dict)
    ocasiones: dict[str, list[tuple[datetime, int]]] = defaultdict(list)

    for log in logs_rows:
        entrega_id = log["entidad_id"]
        items = (log["detalle"] or {}).get("items") or []
        total_ocasion = 0
        for item in items:
            item_id = item.get("id")
            actual = item.get("cantidad_entregada")
            if item_id is None or actual is None:
                continue
            previo = ultimo_conocido[entrega_id].get(item_id, 0)
            delta = actual - previo
            if delta > 0:
                total_ocasion += delta
            ultimo_conocido[entrega_id][item_id] = actual

        if total_ocasion > 0:
            ocasiones[entrega_id].append((log["timestamp"], total_ocasion))

    return ocasiones


def _devoluciones_por_entrega(devoluciones_rows: list) -> dict[str, list[str]]:
    """Un resumen legible por devolucion (ver app/services/devoluciones.py),
    ya ordenadas por fecha -- una entrega puede tener mas de una."""
    resumen: dict[str, list[str]] = defaultdict(list)
    for d in devoluciones_rows:
        entrega_id = str(d["entrega_id"])
        fecha = d["creado_at"].strftime("%d/%m/%Y")
        motivo = _MOTIVOS_LEGIBLES.get(d["motivo"], d["motivo"])
        resolucion = _RESOLUCIONES_LEGIBLES.get(d["resolucion"], d["resolucion"])
        resumen[entrega_id].append(f"{fecha}: {d['cantidad']} und ({motivo}, {resolucion})")
    return resumen


async def generar_reporte_mensual_xlsx(*, sede_id: str | None = None) -> bytes:
    """Una hoja por mes (todo el historico, orden cronologico) -- dentro de
    cada hoja, un bloque de columnas por tipo de documento (Fecha, Número,
    Cantidad entregada, N° entregas, Fechas entregas, Cantidades entregas),
    uno al lado del otro separados por una columna en blanco."""
    pool = await get_pool()
    entregas_rows = await pool.fetch(
        "select id, capturado_at, tipo, indicativo_numero, sede_origen_id from entregas order by capturado_at"
    )
    if sede_id:
        entregas_rows = [r for r in entregas_rows if r["sede_origen_id"] == sede_id]

    ids_relevantes = {str(r["id"]) for r in entregas_rows}
    if not ids_relevantes:
        # Sin entregas (o ninguna de esa sede) -- workbook vacio, sin hojas.
        wb = Workbook()
        wb.remove(wb.active)
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    cantidad_rows = await pool.fetch(
        "select entrega_id, coalesce(sum(cantidad_entregada), 0) as cantidad_total"
        " from entrega_items group by entrega_id"
    )
    cantidad_total_por_entrega = {
        str(r["entrega_id"]): r["cantidad_total"] for r in cantidad_rows if str(r["entrega_id"]) in ids_relevantes
    }

    logs_rows = await pool.fetch(
        "select entidad_id, detalle, \"timestamp\" from logs"
        " where evento = 'entrega_actualizada' order by entidad_id, \"timestamp\""
    )
    ocasiones_por_entrega = _ocasiones_por_entrega(
        [r for r in logs_rows if r["entidad_id"] in ids_relevantes]
    )

    devoluciones_rows = await pool.fetch(
        "select entrega_id, cantidad, motivo, resolucion, creado_at from devoluciones order by entrega_id, creado_at"
    )
    devoluciones_por_entrega = _devoluciones_por_entrega(
        [r for r in devoluciones_rows if str(r["entrega_id"]) in ids_relevantes]
    )

    # (año, mes) -> tipo -> [(fecha, numero, cantidad_total, ocasiones, devoluciones), ...]
    por_mes: dict[
        tuple[int, int], dict[str, list[tuple[datetime, str, int, list[tuple[datetime, int]], list[str]]]]
    ] = defaultdict(lambda: defaultdict(list))
    for r in entregas_rows:
        entrega_id = str(r["id"])
        fecha: datetime = r["capturado_at"]
        tipo = (r["tipo"] or "").strip() or "SIN TIPO"
        cantidad_total = cantidad_total_por_entrega.get(entrega_id, 0)
        ocasiones = sorted(ocasiones_por_entrega.get(entrega_id, []), key=lambda o: o[0])
        devoluciones = devoluciones_por_entrega.get(entrega_id, [])
        por_mes[(fecha.year, fecha.month)][tipo].append(
            (fecha, r["indicativo_numero"] or "", cantidad_total, ocasiones, devoluciones)
        )

    wb = Workbook()
    wb.remove(wb.active)  # la hoja default vacia, cada mes crea la suya

    for (anio, mes) in sorted(por_mes.keys()):
        nombre_hoja = f"{_MESES_ES[mes]} {anio}"[:31]  # Excel limita a 31 caracteres
        ws = wb.create_sheet(title=nombre_hoja)
        tipos_del_mes = por_mes[(anio, mes)]

        columna = 1
        for tipo in sorted(tipos_del_mes.keys(), key=_orden_tipo):
            entradas = sorted(tipos_del_mes[tipo], key=lambda e: e[0])
            columnas_bloque = [get_column_letter(columna + i) for i in range(7)]
            (
                col_fecha, col_numero, col_cantidad, col_n_entregas,
                col_fechas_ent, col_cantidades_ent, col_devoluciones,
            ) = columnas_bloque

            ws.merge_cells(f"{col_fecha}1:{col_devoluciones}1")
            encabezado = ws[f"{col_fecha}1"]
            encabezado.value = tipo
            encabezado.font = Font(bold=True)
            encabezado.alignment = Alignment(horizontal="center")

            titulos = [
                "Fecha", "Número", "Cantidad entregada", "N° entregas",
                "Fechas entregas", "Cantidades entregas", "Devoluciones",
            ]
            for col, titulo in zip(columnas_bloque, titulos):
                celda = ws[f"{col}2"]
                celda.value = titulo
                celda.font = Font(bold=True)

            for i, (fecha, numero, cantidad_total, ocasiones, devoluciones) in enumerate(entradas, start=3):
                ws[f"{col_fecha}{i}"] = fecha.strftime("%d/%m/%Y")
                ws[f"{col_numero}{i}"] = numero
                ws[f"{col_cantidad}{i}"] = cantidad_total
                ws[f"{col_n_entregas}{i}"] = len(ocasiones)
                ws[f"{col_fechas_ent}{i}"] = ", ".join(f.strftime("%d/%m/%Y") for f, _ in ocasiones)
                ws[f"{col_cantidades_ent}{i}"] = ", ".join(str(c) for _, c in ocasiones)
                ws[f"{col_devoluciones}{i}"] = "; ".join(devoluciones)

            ws.column_dimensions[col_fecha].width = 12
            ws.column_dimensions[col_numero].width = 16
            ws.column_dimensions[col_cantidad].width = 16
            ws.column_dimensions[col_n_entregas].width = 11
            ws.column_dimensions[col_fechas_ent].width = 28
            ws.column_dimensions[col_cantidades_ent].width = 22
            ws.column_dimensions[col_devoluciones].width = 40
            columna += 8  # 7 columnas del bloque + 1 en blanco como separador

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
